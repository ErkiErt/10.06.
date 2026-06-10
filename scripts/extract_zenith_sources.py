import json
import math
import re
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl
from pypdf import PdfReader

OUT_DIR = Path("outputs") / "zenith_catalog_database_20260609"
SOURCE_DIR = Path("data_sources")
CATALOG_FILE = SOURCE_DIR / "Zenith catalogue 2020_0.pdf"
PRICE_FILE = SOURCE_DIR / "Pricelist Zenith 2025 v2 (Plastok).xlsx"
PLASTOK_FILE = SOURCE_DIR / "DATABASE.xlsm"
OUT_FILE = OUT_DIR / "source_extract.json"


def clean_text(value):
    if value is None:
        return ""
    text = str(value).strip()
    replacements = {
        "Ãœ": "Ü",
        "Ã•": "Õ",
        "Ã„": "Ä",
        "Ã–": "Ö",
        "Ã¼": "ü",
        "Ãµ": "õ",
        "Ã¤": "ä",
        "Ã¶": "ö",
        "Â°C": "°C",
        "Â°": "°",
    }
    for bad, good in replacements.items():
        text = text.replace(bad, good)
    return text


def value_or_none(value):
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    if isinstance(value, str):
        text = clean_text(value)
        return text if text else None
    return value


def norm_article(value):
    text = clean_text(value).upper()
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def parse_catalog_pdf():
    reader = PdfReader(CATALOG_FILE)
    pages = []
    for index, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""
        pages.append(
            {
                "page": index,
                "text": clean_text(text),
                "char_count": len(text),
            }
        )
    return {
        "file": CATALOG_FILE.name,
        "page_count": len(reader.pages),
        "pages": pages,
    }


def parse_price_list():
    wb = openpyxl.load_workbook(PRICE_FILE, read_only=True, data_only=True)
    ws = wb["ALL"]
    headers = [clean_text(cell) for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    idx = {header: i for i, header in enumerate(headers) if header}
    rows = []
    discounts = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        article = value_or_none(row[idx["Article nr."]])
        if not article or article == "Article nr.":
            continue
        name = value_or_none(row[idx["Name"]])
        if not name:
            continue
        rows.append(
            {
                "article_nr": clean_text(article),
                "article_norm": norm_article(article),
                "name": clean_text(name),
                "gross_price_eur": value_or_none(row[idx["Gross price, EUR"]]),
                "nett_price_eur": value_or_none(row[idx["Nett price, EUR"]]),
                "quantity": value_or_none(row[idx["Quantity"]]),
                "unit": value_or_none(row[idx["Unit"]]),
                "thickness_mm": value_or_none(row[idx["Thickness"]]),
                "width_mm": value_or_none(row[idx["Width"]]),
                "length_mm": value_or_none(row[idx["Length"]]),
                "group": value_or_none(row[idx["Group"]]),
                "material": value_or_none(row[idx["Material"]]),
                "min_temp_c": value_or_none(row[idx["Min °C"]]),
                "max_temp_c": value_or_none(row[idx["Max °C"]]),
                "colour": value_or_none(row[idx["Colour"]]),
                "hardness": value_or_none(row[idx["Hardness"]]),
                "insertion": value_or_none(row[idx["Insertion"]]),
                "insertion_type": value_or_none(row[idx["Insertion type"]]),
                "catalog": value_or_none(row[idx["Catalog"]]),
                "tensile_strength_mpa": value_or_none(row[idx["Tensile strength, Mpa"]]),
                "elongation_pct": value_or_none(row[idx["Elongation, %"]]),
                "source_file": PRICE_FILE.name,
            }
        )

    ws_disc = wb["Discounts"]
    for row in ws_disc.iter_rows(min_row=4, values_only=True):
        if not row or not row[0]:
            continue
        article = clean_text(row[0])
        discount = value_or_none(row[1])
        article_group = value_or_none(row[2])
        if isinstance(discount, (int, float)):
            discounts.append(
                {
                    "article_or_group": article,
                    "discount_pct": discount,
                    "article_group": article_group,
                    "source_file": PRICE_FILE.name,
                }
            )

    grouped = defaultdict(list)
    for row in rows:
        grouped[row["article_norm"]].append(row)

    summary = []
    for article_norm, matches in sorted(grouped.items()):
        gross = [m["gross_price_eur"] for m in matches if isinstance(m["gross_price_eur"], (int, float))]
        nett = [m["nett_price_eur"] for m in matches if isinstance(m["nett_price_eur"], (int, float))]
        summary.append(
            {
                "article_norm": article_norm,
                "article_nr": matches[0]["article_nr"],
                "name": matches[0]["name"],
                "rows": len(matches),
                "unit": "; ".join(sorted({clean_text(m["unit"]) for m in matches if m["unit"]})),
                "gross_min_eur": min(gross) if gross else None,
                "gross_max_eur": max(gross) if gross else None,
                "nett_min_eur": min(nett) if nett else None,
                "nett_max_eur": max(nett) if nett else None,
                "widths_mm": "; ".join(sorted({clean_text(m["width_mm"]) for m in matches if m["width_mm"] is not None}, key=lambda x: (len(x), x))),
                "lengths_mm": "; ".join(sorted({clean_text(m["length_mm"]) for m in matches if m["length_mm"] is not None}, key=lambda x: (len(x), x))),
                "source_file": PRICE_FILE.name,
            }
        )
    return rows, summary, discounts


def parse_plastok_db():
    wb = openpyxl.load_workbook(PLASTOK_FILE, read_only=True, data_only=True, keep_vba=True)
    ws = wb["DATABASE"]
    headers = [clean_text(cell) for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    idx = {header: i for i, header in enumerate(headers) if header}
    wanted = [
        "Article nr.",
        "Name",
        "Nett price, EUR",
        "Gross price, EUR",
        "Quantity",
        "Unit",
        "Category",
        "Thickness",
        "Width",
        "Length",
        "Group",
        "Material",
        "Min °C",
        "Max °C",
        "Colour",
        "Hardness",
        "Insertion",
        "Insertion type",
        "Catalog",
        "Tensile strength, Mpa",
        "Elongation, %",
    ]
    rows = []
    category_counter = Counter()
    group_counter = Counter()
    material_counter = Counter()
    for row in ws.iter_rows(min_row=2, values_only=True):
        article = value_or_none(row[idx["Article nr."]])
        name = value_or_none(row[idx["Name"]])
        if not article and not name:
            continue
        record = {}
        for header in wanted:
            if header in idx:
                record[header] = value_or_none(row[idx[header]])
            else:
                record[header] = None
        rows.append(record)
        category_counter[record.get("Category") or "(puudub)"] += 1
        group_counter[record.get("Group") or "(puudub)"] += 1
        material_counter[record.get("Material") or "(puudub)"] += 1

    summary = []
    combo_counter = Counter()
    for record in rows:
        combo_counter[(record.get("Category") or "(puudub)", record.get("Group") or "(puudub)", record.get("Material") or "(puudub)")] += 1
    for (category, group, material), count in combo_counter.most_common():
        summary.append(
            {
                "category": category,
                "group": group,
                "material": material,
                "rows": count,
                "source_file": PLASTOK_FILE.name,
            }
        )

    return {
        "rows": rows,
        "summary": summary,
        "top_categories": category_counter.most_common(40),
        "top_groups": group_counter.most_common(40),
        "top_materials": material_counter.most_common(80),
    }


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    catalog = parse_catalog_pdf()
    price_rows, price_summary, discounts = parse_price_list()
    plastok = parse_plastok_db()
    payload = {
        "catalog_file": CATALOG_FILE.name,
        "catalog_page_count": catalog["page_count"],
        "catalog_pages": catalog["pages"],
        "price_file": PRICE_FILE.name,
        "plastok_file": PLASTOK_FILE.name,
        "price_rows": price_rows,
        "price_summary": price_summary,
        "discounts": discounts,
        "plastok_rows": plastok["rows"],
        "plastok_summary": plastok["summary"],
        "plastok_top_categories": plastok["top_categories"],
        "plastok_top_groups": plastok["top_groups"],
        "plastok_top_materials": plastok["top_materials"],
    }
    OUT_FILE.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(
        json.dumps(
            {
                "out": str(OUT_FILE),
                "catalog_pages": catalog["page_count"],
                "price_rows": len(price_rows),
                "price_summary": len(price_summary),
                "discounts": len(discounts),
                "plastok_rows": len(plastok["rows"]),
                "plastok_summary": len(plastok["summary"]),
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
